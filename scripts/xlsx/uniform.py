# -*- coding: utf-8 -*-
"""Κάθε πίνακας ζυγισμένος: δεκάδες κάτω από δεκάδες.

ΤΙ ΣΗΜΑΙΝΕΙ «ΣΤΟΙΧΙΣΜΕΝΟ» ΣΕ ΦΥΛΛΟ EXCEL. Τα ποσά γράφονται ως κείμενο για να
κρατήσουν την ελληνική μορφή σε κάθε υπολογιστή. Κείμενο στοιχισμένο δεξιά
ευθυγραμμίζεται από το ΤΕΛΟΣ του: «650,00 €» κάτω από «9.600,00 €» δίνει
μονάδες κάτω από μονάδες ΜΟΝΟ αν και τα δύο έχουν τα ίδια δεκαδικά και την ίδια
μονάδα. Αρκεί ένα «85,5» δίπλα σε ένα «120,00» και η στήλη γέρνει.

Ο έλεγχος κοιτά κάθε ΠΙΝΑΚΑ χωριστά και όχι κάθε στήλη του φύλλου: δύο πίνακες
στο ίδιο φύλλο έχουν διαφορετικές στήλες και δεν οφείλουν να συμφωνούν.
Ζητούνται τέσσερα, ανά στήλη ενός πίνακα:

  · μία στοίχιση, ίδια σε όλες τις γραμμές
  · τα ίδια δεκαδικά ψηφία για την ίδια μονάδα
  · η ίδια μονάδα (€, %, τίποτα)
  · καμία γραμμή χωρίς πλαίσιο ανάμεσα σε γραμμές που έχουν

ΔΥΟ ΕΞΑΙΡΕΣΕΙΣ, ΓΡΑΜΜΕΝΕΣ ΚΑΙ ΟΧΙ ΣΙΩΠΗΛΕΣ:

  · Η ΣΗΜΕΙΩΣΗ ΔΕΝ ΕΙΝΑΙ ΓΡΑΜΜΗ ΠΙΝΑΚΑ. Οι προτάσεις κάτω από έναν πίνακα
    απλώνονται σε όλο το πλάτος και δεν έχουν πλαίσιο, εξ ορισμού. Ξεχωρίζουν
    γιατί έχουν περιεχόμενο ΜΟΝΟ στην πρώτη στήλη.
  · ΚΑΤΑΛΟΓΟΣ «ΠΑΡΑΜΕΤΡΟΣ / ΤΙΜΗ» ΔΕΝ ΕΙΝΑΙ ΣΤΗΛΗ ΜΙΑΣ ΜΕΤΡΗΣΗΣ. «Δεκάδες
    κάτω από δεκάδες» έχει νόημα όταν η στήλη μετράει το ΙΔΙΟ πράγμα. Σε
    κατάλογο ρυθμίσεων η μία γραμμή είναι ευρώ, η επόμενη ποσοστό και η τρίτη
    νύχτες: η μονάδα επιτρέπεται να αλλάζει, τα δεκαδικά της ΙΔΙΑΣ μονάδας όχι.
"""
import sys, os, glob, re, collections
import openpyxl

HEADBG = 'E9ECEF'
QUANTITY = re.compile(r'^(-?[\d.]*\d(?:,\d+)?)\s*(€|%)$')
DATE = re.compile(r'^(dd|mm|yyyy|/|\.|-)+$')


def is_head(cell):
    f = cell.fill
    return (f is not None and f.fgColor is not None
            and str(f.fgColor.rgb or '').upper().endswith(HEADBG))


# Μορφή του Excel με δεκαδικά «όσα χρειάζονται»: «#,##0.##». Μια τέτοια στήλη
# δείχνει «78» δίπλα σε «85,5» και δεν στοιχίζεται ποτέ, όσο δεξιά κι αν είναι.
VARIABLE = re.compile(r'0\.#')


def shape(cell):
    """(δεκαδικά, μονάδα) για ποσό ή ποσοστό. None για ο,τιδήποτε άλλο."""
    v = cell.value
    if isinstance(v, (int, float)):
        fmt = cell.number_format or ''
        if fmt in ('General', '@') or DATE.match(fmt):
            return None
        if VARIABLE.search(fmt):
            return ('μεταβλητά', 'αριθμός')
        m = re.search(r'0\.(0+)', fmt)
        return ((len(m.group(1)) if m else 0), 'αριθμός')
    m = QUANTITY.match(str(v or '').strip())
    if not m:
        return None
    num, unit = m.groups()
    return (len(num.split(',')[1]) if ',' in num else 0, unit)


def note_row(ws, r, width):
    """Σημείωση κάτω από πίνακα: περιεχόμενο μόνο στην πρώτη στήλη."""
    filled = [c for c in range(1, width + 1) if ws.cell(row=r, column=c).value not in (None, '')]
    return filled == [1]


def blocks(ws):
    """Κάθε επικεφαλίδα και οι γραμμές της, ώς την πρώτη κενή ή την επόμενη."""
    heads = [r for r in range(1, ws.max_row + 1)
             if sum(1 for c in range(1, ws.max_column + 1) if is_head(ws.cell(row=r, column=c))) >= 2]
    for i, hr in enumerate(heads):
        stop = heads[i + 1] if i + 1 < len(heads) else ws.max_row + 1
        width = max((c for c in range(1, ws.max_column + 1)
                     if is_head(ws.cell(row=hr, column=c))), default=0)
        rows = []
        for r in range(hr + 1, stop):
            if all(ws.cell(row=r, column=c).value in (None, '') for c in range(1, width + 1)):
                break
            rows.append(r)
        if rows:
            yield hr, rows, width


def check(path):
    wb = openpyxl.load_workbook(path)
    name = os.path.basename(path)
    problems = []
    for ws in wb.worksheets:
        for hr, rows, width in blocks(ws):
            # Η ΕΠΙΚΕΦΑΛΙΔΑ ΕΙΝΑΙ ΜΙΑ ΓΡΑΜΜΗ, ΟΧΙ ΜΙΣΗ.
            aligns = {ws.cell(row=hr, column=c).alignment.horizontal for c in range(1, width + 1)}
            if len(aligns) > 1:
                problems.append('%s γρ.%d: η επικεφαλίδα έχει %s' % (ws.title, hr, sorted(map(str, aligns))))
            # Κατάλογος «Παράμετρος / Τιμή»: δύο στήλες και η δεύτερη λέγεται «Τιμή».
            keyvalue = width == 2 and str(ws.cell(row=hr, column=2).value or '').strip() == 'Τιμή'
            data = [r for r in rows if not note_row(ws, r, width)]
            for c in range(1, width + 1):
                seen, aligned, unstyled = set(), set(), []
                for r in data:
                    cell = ws.cell(row=r, column=c)
                    if cell.value in (None, ''):
                        continue
                    sh = shape(cell)
                    if sh:
                        seen.add(sh)
                        aligned.add(cell.alignment.horizontal)
                    if cell.border.bottom.style is None and cell.border.top.style is None:
                        unstyled.append(r)
                col = ws.cell(row=hr, column=c).value or ('στήλη %d' % c)
                if any(d == 'μεταβλητά' for d, _ in seen):
                    problems.append('%s / «%s»: μορφή με μεταβλητά δεκαδικά' % (ws.title, col))
                    seen = {s for s in seen if s[0] != 'μεταβλητά'}
                decimals = collections.defaultdict(set)
                for dec, unit in seen:
                    decimals[unit].add(dec)
                ragged = {u: sorted(d) for u, d in decimals.items() if len(d) > 1}
                if ragged:
                    problems.append('%s / «%s»: άνισα δεκαδικά %s'
                                    % (ws.title, col, ragged))
                elif len(seen) > 1 and not keyvalue:
                    problems.append('%s / «%s»: ανάμεικτες μονάδες %s'
                                    % (ws.title, col, sorted(u for _, u in seen)))
                if aligned - {'right'}:
                    problems.append('%s / «%s»: ποσά με στοίχιση %s'
                                    % (ws.title, col, sorted(map(str, aligned))))
                if unstyled and len(unstyled) < len(data):
                    problems.append('%s / «%s»: γραμμές χωρίς πλαίσιο %s'
                                    % (ws.title, col, unstyled[:4]))
    return name, problems


def main(out):
    fails = 0
    for path in sorted(glob.glob(os.path.join(out, '*.xlsx'))):
        name, problems = check(path)
        if problems:
            fails += len(problems)
            print('  ✗ %s' % name)
            for p in problems:
                print('      %s' % p)
        else:
            print('  ✓ %s: κάθε πίνακας ζυγισμένος' % name)
    return 1 if fails else 0


sys.exit(main(sys.argv[1]))
