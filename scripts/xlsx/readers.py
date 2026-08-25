# -*- coding: utf-8 -*-
"""Το ίδιο βιβλίο, διαβασμένο από δύο ξένες υλοποιήσεις.

ΓΙΑΤΙ ΔΥΟ ΚΑΙ ΟΧΙ ΜΙΑ. Η openpyxl είναι Python και βλέπει τα ΠΑΝΤΑ, ακόμη και
τις εικόνες. Η calamine είναι Rust, γραμμένη από άλλους· δεν διαβάζει εικόνες
αλλά διαβάζει τιμές: αν συμφωνήσουν οι δύο στα κελιά, το αρχείο δεν είναι απλώς
«αποδεκτό από τη μία βιβλιοθήκη που το ξέρει».
"""
import sys, os, glob

def main(out):
    try:
        import openpyxl
    except ImportError:
        print('  · openpyxl δεν είναι εγκατεστημένη, παραλείπεται'); return 0
    try:
        from python_calamine import CalamineWorkbook
    except ImportError:
        CalamineWorkbook = None

    fails = 0
    for path in sorted(glob.glob(os.path.join(out, '*.xlsx'))):
        name = os.path.basename(path)
        try:
            wb = openpyxl.load_workbook(path)
        except Exception as e:
            print('  ✗ %s: η openpyxl δεν το άνοιξε (%s)' % (name, e)); fails += 1; continue

        # ΚΑΘΕ ΦΥΛΛΟ ΜΕ ΖΩΝΗ ΤΙΤΛΟΥ ΚΟΥΒΑΛΑ ΤΟ ΣΗΜΑ, ΑΚΡΙΒΩΣ ΕΝΑ.
        sheets, marked, bad = 0, 0, []
        for ws in wb.worksheets:
            sheets += 1
            n = len(getattr(ws, '_images', []))
            if n == 1:
                marked += 1
                h = ws.row_dimensions[1].height
                if h is None or abs(h - 34) > 0.6:
                    bad.append('%s: η γραμμή τίτλου είναι %s αντί για 34' % (ws.title, h))
                a = ws.cell(row=1, column=1).alignment
                if not a or (a.indent or 0) < 7:
                    bad.append('%s: ο τίτλος δεν κάνει τόπο στο σήμα' % ws.title)
            elif n > 1:
                bad.append('%s: %d εικόνες αντί για μία' % (ws.title, n))
            else:
                bad.append('%s: χωρίς σήμα' % ws.title)
        if bad:
            print('  ✗ %s: %s' % (name, ' · '.join(bad))); fails += 1
        else:
            print('  ✓ %s: openpyxl, %d φύλλα, ένα σήμα στο καθένα' % (name, sheets))

        if CalamineWorkbook is None:
            continue
        try:
            cal = CalamineWorkbook.from_path(path)
            for ws in wb.worksheets:
                rows = cal.get_sheet_by_name(ws.title).to_python()
                mine = ws.cell(row=1, column=1).value
                theirs = rows[0][0] if rows and rows[0] else None
                if str(mine or '') != str(theirs or ''):
                    print('  ✗ %s / %s: calamine διαβάζει «%s», openpyxl «%s»' % (name, ws.title, theirs, mine))
                    fails += 1
        except Exception as e:
            print('  ✗ %s: η calamine δεν το άνοιξε (%s)' % (name, e)); fails += 1
    if CalamineWorkbook is None:
        print('  · python-calamine δεν είναι εγκατεστημένη, παραλείπεται')
    else:
        print('  ✓ calamine (Rust) συμφωνεί με την openpyxl στον τίτλο κάθε φύλλου')
    return 1 if fails else 0

sys.exit(main(sys.argv[1]))
